# taxasarbi.py — V13.3 + run_taxas headless (com export formatado)
# - FINAL só com linhas OK (sem falha de base)
# - PENDENTES só com linhas que falharam lookup de base (campos do FUNDO)
# - Vazio trata "nan"/"none"/"null"/"na"/"nat"
# - VALOR_FAVORECIDO mantido como TEXTO com '4444' durante o fluxo e
#   convertido para número (6 casas) somente no export
# - DATA_PAGAMENTO exportada como TEXTO

from __future__ import annotations
import sys, re, unicodedata
from pathlib import Path
import pandas as pd
from decimal import Decimal, InvalidOperation
from tkinter import Tk, filedialog, messagebox
import datetime as dt

# ---------------- utils ----------------
def resource_path(rel: str) -> Path:
    exe_dir = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    p = exe_dir / rel
    if p.exists():
        return p
    try:
        return Path(sys._MEIPASS) / rel  # type: ignore[attr-defined]
    except Exception:
        return Path(__file__).parent / rel

def normalize_str(s: str) -> str:
    s = str(s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s)

def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s or ""))

def conta_key_strip0(s: str) -> str:
    d = only_digits(s)
    d2 = d.lstrip("0")
    return d2 if d2 != "" else "0"

def conta_key_last8(s: str) -> str:
    d = only_digits(s)
    return d[-8:] if len(d) >= 8 else d

def clean_text(s: str) -> str:
    s = str(s or "").upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    return re.sub(r"\s+", " ", s).strip()

def parse_decimal_like(s: str) -> Decimal:
    s = str(s or "").strip()
    if s == "":
        raise InvalidOperation("vazio")
    if "," in s and (s.count(".") == 0 or s.rfind(",") > s.rfind(".")):
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    return Decimal(s)

def valor_4444(v):
    """Gera string padrão com duas casas + '4444' (ex.: 18645,97 -> '18645.974444')."""
    try:
        d = parse_decimal_like(v)
        return f"{d:.2f}4444"
    except Exception:
        return ""

# valida data dd/mm/aaaa e não aceita datas passadas
def valida_data_pagamento(s: str) -> str:
    s = re.sub(r"[.\- ]", "/", str(s).strip())
    hoje = dt.date.today()
    try:
        d = dt.datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        raise ValueError("Use o formato dd/mm/aaaa (ex.: 23/07/2025).")
    if d < hoje:
        raise ValueError(
            f"A data informada ({d.strftime('%d/%m/%Y')}) é anterior à data atual ({hoje.strftime('%d/%m/%Y')})."
        )
    return d.strftime("%d/%m/%Y")

# Tokens que devem ser tratados como vazio
EMPTY_TOKENS = {"", "nan", "none", "null", "na", "n/a", "nat"}

def is_blank_series(s: pd.Series) -> pd.Series:
    return s.isna() | s.astype(str).str.strip().str.lower().isin(EMPTY_TOKENS)

def sanitize_df_na(df: pd.DataFrame) -> pd.DataFrame:
    return df.replace({r"^\s*(nan|none|null|na|n/a|nat)\s*$": ""}, regex=True).fillna("")

# --- FIXOS do favorecido (sem CONTA/VALOR para não sobrescrever o usuário) ---
CONSTANTES_FAV = {
    "CODIGO_BANCO_FAVORECIDO": "213",
    "AGENCIA_FAVORECIDO": "0001",
    "LOGRADOURO_FAVORECIDO": "AV PRES JUSCELINO KUBITSCHEK",
    "NUMERO_FAVORECIDO": "1726",
    "COMPLEMENTO_FAVORECIDO": "CJ 92",
    "BAIRRO_FAVORECIDO": "ITAIM BIBI",
    "CEP_FAVORECIDO": "04543000",
    "CIDADE_FAVORECIDO": "SAO PAULO",
    "UF_FAVORECIDO": "SP",
}

# ---------------- colunas finais ----------------
COLUNAS_FINAIS = [
    "CNPJ_FUNDO","CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO",
    "NOME_FAVORECIDO","CADASTRO_FAVORECIDO",
    "CODIGO_BANCO_FAVORECIDO","AGENCIA_FAVORECIDO","CONTA_COM_DIGITO_FAVORECIDO",
    "VALOR_FAVORECIDO","LOGRADOURO_FAVORECIDO","NUMERO_FAVORECIDO","COMPLEMENTO_FAVORECIDO",
    "BAIRRO_FAVORECIDO","CEP_FAVORECIDO","CIDADE_FAVORECIDO","UF_FAVORECIDO","DATA_PAGAMENTO"
]

# ---------------- base ----------------
def carregar_base(caminho: Path) -> pd.DataFrame:
    base = pd.read_excel(caminho, dtype=str, engine="openpyxl")
    base.columns = [normalize_str(c) for c in base.columns]

    req = ["CONTA_COM_DIGITO_FAVORECIDO", "CNPJ_FUNDO", "NOME_FAVORECIDO"]
    for c in req:
        if c not in base.columns:
            raise ValueError(f"A base precisa da coluna '{c}'")

    for k in ["CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO"]:
        if k not in base.columns:
            base[k] = ""

    base["__K_STRIP0__"] = base["CONTA_COM_DIGITO_FAVORECIDO"].map(conta_key_strip0)
    base["__K_LAST8__"]  = base["CONTA_COM_DIGITO_FAVORECIDO"].map(conta_key_last8)

    base["CNPJ_FUNDO"]             = base["CNPJ_FUNDO"].astype(str)
    base["NOME_FAVORECIDO"]        = base["NOME_FAVORECIDO"].map(clean_text)
    base["CODIGO_BANCO_FUNDO"]     = base["CODIGO_BANCO_FUNDO"].map(only_digits)
    base["AGENCIA_FUNDO"]          = base["AGENCIA_FUNDO"].map(only_digits)
    base["CONTA_COM_DIGITO_FUNDO"] = base["CONTA_COM_DIGITO_FUNDO"].astype(str).str.replace(r"\s+","",regex=True)

    base = sanitize_df_na(base)
    base = base[base["__K_STRIP0__"] != ""].drop_duplicates("__K_STRIP0__", keep="first").copy()
    return base[["__K_STRIP0__","__K_LAST8__","CNPJ_FUNDO","NOME_FAVORECIDO",
                 "CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO"]]

# ---------------- usuário ----------------
SINAIS_CONTA_FAV = ["NROCONTA","NRO CONTA","NUMERO DA CONTA","CONTA","CONTA FAVORECIDO"]
SINAIS_VALOR     = ["VALOR","TARIFA","VALOR PAGO","VALOR A PAGAR"]
SINAIS_BANCO_FAV = ["CODIGO BANCO","COD BANCO","BANCO FAVORECIDO","CODIGO_BANCO","BANCO"]
SINAIS_AG_FAV    = ["AGENCIA","AGÊNCIA","AGENCIA FAVORECIDO","AG"]
SINAIS_DATA      = ["DATA","PAGAMENTO","DATA PAGAMENTO","DT_PAGAMENTO"]

def achar_col_usuario(df: pd.DataFrame, preferida: str | None, sinais: list[str]) -> str | None:
    mapa = {normalize_str(c).upper().replace(":", ""): c for c in df.columns}
    if preferida:
        p = normalize_str(preferida).upper().replace(":", "")
        if p in mapa:
            return mapa[p]
    for c in df.columns:
        u = normalize_str(c).upper().replace(":", "")
        if any(sig in u for sig in sinais):
            return c
    return None

def ler_planilha_usuario(p: Path) -> pd.DataFrame:
    if p.suffix.lower() == ".csv":
        df = pd.read_csv(p, dtype=str, sep=None, engine="python")
    else:
        df = pd.read_excel(p, dtype=str, engine="openpyxl")
    df.columns = [normalize_str(c) for c in df.columns]
    return sanitize_df_na(df)

# ---------------- seleção de arquivo do usuário ----------------
def selecionar_arquivo_usuario() -> Path:
    Tk().withdraw()
    p = filedialog.askopenfilename(
        title="Selecione a planilha do usuário",
        filetypes=[("Planilhas Excel","*.xlsx;*.xls"),("CSV","*.csv"),("Todos","*.*")],
    )
    if not p:
        raise RuntimeError("Nenhum arquivo selecionado.")
    return Path(p)

# ---------------- construção final ----------------
def construir_final(df_user: pd.DataFrame, base: pd.DataFrame) -> pd.DataFrame:
    col_conta = achar_col_usuario(df_user, "NROCONTA:", SINAIS_CONTA_FAV)
    col_valor = achar_col_usuario(df_user, "Valor:",    SINAIS_VALOR)
    col_bco_f = achar_col_usuario(df_user, None,        SINAIS_BANCO_FAV)
    col_ag_f  = achar_col_usuario(df_user, None,        SINAIS_AG_FAV)
    col_data  = achar_col_usuario(df_user, None,        SINAIS_DATA)

    if not col_conta: raise ValueError("Não encontrei a coluna 'NROCONTA:' na planilha do usuário.")
    if not col_valor: raise ValueError("Não encontrei a coluna 'Valor:' na planilha do usuário.")

    w = df_user.copy()
    w["__K_STRIP0__"] = w[col_conta].map(conta_key_strip0)
    w["__K_LAST8__"]  = w[col_conta].map(conta_key_last8)

    out = w.merge(base, on="__K_STRIP0__", how="left", suffixes=("", "_B"))

    falt = out["CNPJ_FUNDO"].isna() | out["CNPJ_FUNDO"].astype(str).str.strip().eq("")
    if falt.any():
        aux = base.set_index("__K_LAST8__")
        for col in ["CNPJ_FUNDO","NOME_FAVORECIDO","CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO"]:
            out.loc[falt, col] = out.loc[falt, "__K_LAST8__"].map(aux[col])

    final = pd.DataFrame(index=out.index, columns=COLUNAS_FINAIS).fillna("")
    final["CNPJ_FUNDO"]             = out["CNPJ_FUNDO"]
    final["NOME_FAVORECIDO"]        = out["NOME_FAVORECIDO"].map(clean_text)
    final["CODIGO_BANCO_FUNDO"]     = out["CODIGO_BANCO_FUNDO"]
    final["AGENCIA_FUNDO"]          = out["AGENCIA_FUNDO"]
    final["CONTA_COM_DIGITO_FUNDO"] = out["CONTA_COM_DIGITO_FUNDO"]

    final["CNPJ_FUNDO"] = final["CNPJ_FUNDO"].where(~is_blank_series(final["CNPJ_FUNDO"]),
                                                    "FUNDO NAO CADASTRADO NO BANCO DE DADOS")

    # <- VALOR_FAVORECIDO permanece TEXTO com '4444' até o export
    final["CONTA_COM_DIGITO_FAVORECIDO"] = w[col_conta].astype(str)
    final["VALOR_FAVORECIDO"]            = w[col_valor].apply(valor_4444).astype(str)
    final["CADASTRO_FAVORECIDO"]         = final["CNPJ_FUNDO"]
    if col_bco_f: final["CODIGO_BANCO_FAVORECIDO"] = w[col_bco_f].map(only_digits)
    if col_ag_f:  final["AGENCIA_FAVORECIDO"]      = w[col_ag_f].map(only_digits)
    if col_data:  final["DATA_PAGAMENTO"]          = w[col_data].astype(str)

    # Aplica FIXOS restantes (sem sobrescrever conta/valor do usuário)
    for k, v in CONSTANTES_FAV.items():
        final[k] = v

    return sanitize_df_na(final)

# ---------------- pendências ----------------
def mask_pendencias(df_final: pd.DataFrame) -> pd.Series:
    nao_cadastrado = df_final["CNPJ_FUNDO"].astype(str).str.contains("FUNDO NAO CADASTRADO", na=False)
    faltam_fundos = (
        is_blank_series(df_final["CODIGO_BANCO_FUNDO"]) |
        is_blank_series(df_final["AGENCIA_FUNDO"]) |
        is_blank_series(df_final["CONTA_COM_DIGITO_FUNDO"])
    )
    return nao_cadastrado | faltam_fundos

def split_final_pendentes(df_final: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    m = mask_pendencias(df_final)
    df_pend = df_final.loc[m].copy()
    df_ok   = df_final.loc[~m].copy()
    return sanitize_df_na(df_ok), sanitize_df_na(df_pend)

# ---------------- helpers de exportação ----------------
def _valor_str_to_float6(v: str):
    """
    Converte strings como '18645.974444' ou '18.645,97' para float com 6 casas.
    Mantém o sufixo '4444' quando existir. Retorna None se não der para converter.
    """
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except Exception:
        return None

def _parse_date_flex(val):
    """
    Converte vários formatos para datetime:
    - 'dd/mm/aaaa', 'dd/mm/aa', 'aaaa-mm-dd'
    - número serial do Excel
    Retorna None se não converter.
    """
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    # formatos comuns
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    # serial Excel
    try:
        sn = re.sub(r"[^0-9.\-]", "", s)
        if sn != "":
            serial = float(sn)
            base = dt.datetime(1899, 12, 30)
            d = base + dt.timedelta(days=serial)
            return dt.datetime(d.year, d.month, d.day)
    except Exception:
        pass
    return None

def _write_excel_formatted(df: pd.DataFrame, dest: Path, sheet_name: str) -> Path:
    """Grava Excel: todas colunas TEXTO, exceto VALOR_FAVORECIDO (0.000000). DATA_PAGAMENTO sai como TEXTO."""
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        cols = list(df.columns)
        n_rows, n_cols = df.shape
        idx_val = cols.index("VALOR_FAVORECIDO") + 1 if "VALOR_FAVORECIDO" in cols else None
        idx_dt  = cols.index("DATA_PAGAMENTO") + 1 if "DATA_PAGAMENTO" in cols else None

        # Cabeçalho como texto
        for c in range(1, n_cols + 1):
            ws.cell(row=1, column=c).number_format = "@"

        # Linhas: tudo texto; VALOR como número; DATA_PAGAMENTO forçada TEXTO
        for r in range(2, n_rows + 2):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)

                if idx_val and c == idx_val:
                    f = _valor_str_to_float6(cell.value)
                    cell.value = f if f is not None else None
                    cell.number_format = "0.000000"

                elif idx_dt and c == idx_dt:
                    # >>> ALTERADO: manter DATA_PAGAMENTO como TEXTO <<<
                    cell.value = "" if cell.value is None else str(cell.value)
                    cell.number_format = "@"

                else:
                    cell.value = "" if cell.value is None else str(cell.value)
                    cell.number_format = "@"

        # Larguras
        for c in range(1, n_cols + 1):
            colname = cols[c-1]
            if idx_val and c == idx_val:
                width = 18
            else:
                maxlen = max(len(str(colname)), *(len(str(x)) for x in df[colname].astype(str).values)) + 2
                width = min(maxlen, 60)
            ws.column_dimensions[get_column_letter(c)].width = width

    return dest

# ---------------- export ----------------
def exportar(df: pd.DataFrame, origem: Path, destino_forcado: Path | None = None) -> Path:
    dest = Path(destino_forcado) if destino_forcado else origem.with_name(origem.stem + "_TAXAS_ARBI_FINAL.xlsx")
    return _write_excel_formatted(df, dest, "TAXAS_ARBI")

def exportar_pendentes(df: pd.DataFrame, origem: Path | None = None, destino_final: Path | None = None) -> Path:
    # Padrão: TaxasArbi_Pendentes_AAAAMMDD
    if destino_final:
        final = Path(destino_final)
        data_hoje = dt.date.today().strftime('%Y%m%d')
        pend = final.parent / f"TaxasArbi_Pendentes_{data_hoje}.xlsx"
    else:
        data_hoje = dt.date.today().strftime('%Y%m%d')
        pend = Path(origem).parent / f"TaxasArbi_Pendentes_{data_hoje}.xlsx"
    return _write_excel_formatted(df, pend, "PENDENTES")

# ---------------- HEADLESS (para launcher) ----------------
def run_taxas(input_file: str | Path, data_pagamento: str, saida_path: str | Path | None = None) -> dict:
    data_ok = valida_data_pagamento(data_pagamento)

    base_path = resource_path("Basedadosfundos_Arbi.xlsx")
    base = carregar_base(base_path)

    in_path = Path(input_file)
    df_user = ler_planilha_usuario(in_path)

    df_final_all = construir_final(df_user, base)
    df_final_all["DATA_PAGAMENTO"] = data_ok  # manter como string; export será TEXTO

    df_final_ok, df_pend = split_final_pendentes(df_final_all)

    # Padrão: Processo_Tipo_AAAAMMDD
    if saida_path:
        out_final = Path(saida_path)
    else:
        data_hoje = dt.date.today().strftime('%Y%m%d')
        out_final = in_path.parent / f"TaxasArbi_Final_{data_hoje}.xlsx"
    
    caminho_final = exportar(df_final_ok, origem=in_path, destino_forcado=out_final)

    caminho_pend = None
    if not df_pend.empty:
        caminho_pend = exportar_pendentes(df_pend, destino_final=caminho_final)

    return {
        "saida_taxas_final": str(caminho_final.resolve()),
        "saida_taxas_pendentes": str(Path(caminho_pend).resolve()) if caminho_pend else "",
        "qtd_total": int(len(df_final_all)),
        "qtd_ok": int(len(df_final_ok)),
        "qtd_pendentes": int(len(df_pend)),
    }
