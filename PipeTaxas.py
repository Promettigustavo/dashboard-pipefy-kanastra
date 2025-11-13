# pipetaxas.py
from __future__ import annotations
import pandas as pd
import numpy as np
from pathlib import Path
import re, unicodedata
import datetime as dt

# -------- Utils --------
def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))

def digits_only(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)): return ""
    return re.sub(r"\D", "", str(x))

def clean_text_keep_basic(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)): return ""
    s = strip_accents(str(x)).upper().strip()
    s = re.sub(r"[.,;:!?(){}\[\]'\"`´^~•·º°ª]", " ", s)  # preserva / & -
    return re.sub(r"\s+", " ", s).strip()

def normalize_agencia(x) -> str:
    """
    3–4 dígitos; se vier 5:
      - começa com '0' -> remove o 1º
      - começa diferente de '1' -> remove o último
      - senão -> pega os 4 primeiros
    """
    ag = digits_only(x)
    if len(ag) == 5:
        if ag.startswith("0"):
            ag = ag[1:]
        elif not ag.startswith("1"):
            ag = ag[:-1]
        else:
            ag = ag[:4]
    return ag

def normalize_conta(x) -> str:
    return digits_only(x)

def parse_money_to_float(x):
    if x is None or (isinstance(x, float) and np.isnan(x)) or str(x).strip() == "":
        return np.nan
    s = str(x).strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except:
        return np.nan

def money_with_4444(x):
    """
    1) normaliza p/ 2 casas
    2) concatena '4444'
    3) retorna número (float) -> ex.: 15638,91 -> 15638.914444
    """
    v = parse_money_to_float(x)
    if pd.isna(v):
        return np.nan
    s = f"{v:.2f}"
    try:
        return float(s + "4444")
    except:
        return np.nan

def to_excel_date(dmy: str):
    # >>> ALTERADO: retorna TEXTO 'dd/mm/yyyy' (antes retornava Timestamp)
    ts = pd.to_datetime(dmy, dayfirst=True, errors="coerce")
    return "" if pd.isna(ts) else ts.strftime("%d/%m/%Y")

def read_user_table(path: Path) -> pd.DataFrame:
    path = Path(path)
    if path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        return pd.read_excel(path, dtype=str)
    try:
        return pd.read_csv(path, dtype=str, sep=";")
    except:
        return pd.read_csv(path, dtype=str, sep=",")

# -------- Layout --------
COLS = [
    "CNPJ_FUNDO","CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO",
    "NOME_FAVORECIDO","CADASTRO_FAVORECIDO",
    "CODIGO_BANCO_FAVORECIDO","AGENCIA_FAVORECIDO","CONTA_COM_DIGITO_FAVORECIDO",
    "VALOR_FAVORECIDO","LOGRADOURO_FAVORECIDO","NUMERO_FAVORECIDO",
    "COMPLEMENTO_FAVORECIDO","BAIRRO_FAVORECIDO","CEP_FAVORECIDO",
    "CIDADE_FAVORECIDO","UF_FAVORECIDO","DATA_PAGAMENTO",
]

# -------- Core --------
def _normalize(df: pd.DataFrame, data_pagamento: str) -> pd.DataFrame:
    # garante colunas
    for c in COLS:
        if c not in df.columns:
            df[c] = np.nan

    # CODIGO_BANCO_* -> só números
    for c in ["CODIGO_BANCO_FUNDO", "CODIGO_BANCO_FAVORECIDO"]:
        df[c] = df[c].map(digits_only)

    # textos
    for c in ["NOME_FAVORECIDO","LOGRADOURO_FAVORECIDO","NUMERO_FAVORECIDO",
              "COMPLEMENTO_FAVORECIDO","BAIRRO_FAVORECIDO","CIDADE_FAVORECIDO",
              "UF_FAVORECIDO"]:
        df[c] = df[c].map(clean_text_keep_basic)

    # dígitos
    df["CNPJ_FUNDO"] = df["CNPJ_FUNDO"].map(digits_only)
    df["CADASTRO_FAVORECIDO"] = df["CADASTRO_FAVORECIDO"].map(digits_only)
    df["CEP_FAVORECIDO"] = df["CEP_FAVORECIDO"].map(digits_only)

    # agência/conta
    for c in ["AGENCIA_FUNDO","AGENCIA_FAVORECIDO"]:
        df[c] = df[c].map(normalize_agencia)
    for c in ["CONTA_COM_DIGITO_FUNDO","CONTA_COM_DIGITO_FAVORECIDO"]:
        df[c] = df[c].map(normalize_conta)

    # valor e data
    df["VALOR_FAVORECIDO"] = df["VALOR_FAVORECIDO"].map(money_with_4444)   # número
    df["DATA_PAGAMENTO"]   = to_excel_date(data_pagamento)                 # >>> TEXTO 'dd/mm/yyyy'

    # tudo TEXTO exceto valor (data também fica TEXTO)
    for c in COLS:
        if c != "VALOR_FAVORECIDO":  # <<< DATA_PAGAMENTO entra como texto aqui
            df[c] = df[c].fillna("").astype(str)
    
    # FORÇAR campos de endereço com valores fixos (SEMPRE sobrescrever)
    df["LOGRADOURO_FAVORECIDO"] = "AVENIDA DOS VINHEDOS"
    df["NUMERO_FAVORECIDO"] = "71"
    df["COMPLEMENTO_FAVORECIDO"] = "SALA 802"
    df["BAIRRO_FAVORECIDO"] = "JARDIM SUL"
    df["CEP_FAVORECIDO"] = "38411848"
    df["CIDADE_FAVORECIDO"] = "UBERLANDIA"
    df["UF_FAVORECIDO"] = "MG"

    return df[COLS].copy()

def _save_with_formats(df: pd.DataFrame, path_xlsx: Path,
                       number_format: str = "0.000000",
                       date_format: str = "dd/mm/yyyy",
                       sheet_name: str = "FINAL") -> None:
    """Força formatos no Excel: todas as colunas TEXTO, exceto VALOR como número.
       DATA_PAGAMENTO agora permanece TEXTO.
    """
    text_cols = [c for c in df.columns if c != "VALOR_FAVORECIDO"]
    val_idx = df.columns.get_loc("VALOR_FAVORECIDO")

    # 1) xlsxwriter
    try:
        with pd.ExcelWriter(path_xlsx, engine="xlsxwriter",
                            date_format=date_format, datetime_format=date_format) as w:
            df.to_excel(w, index=False, sheet_name=sheet_name)
            ws = w.sheets[sheet_name]; wb = w.book
            fmt_text = wb.add_format({'num_format': '@'})
            fmt_val  = wb.add_format({'num_format': number_format})
            # texto (inclui DATA_PAGAMENTO)
            for c in text_cols:
                i = df.columns.get_loc(c)
                ws.set_column(i, i, None, fmt_text)
            # valor
            ws.set_column(val_idx, val_idx, None, fmt_val)
        return
    except Exception:
        pass

    # 2) Fallback: openpyxl
    with pd.ExcelWriter(path_xlsx, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path_xlsx)
        ws = wb[sheet_name]

        def _col_letter(idx0: int) -> str:
            n = idx0; s = ""
            while n:
                n, r = divmod(n-1, 26)
                s = chr(65 + r) + s
            return s

        # texto (inclui DATA_PAGAMENTO)
        for c in text_cols:
            i = df.columns.get_loc(c) + 1
            col = _col_letter(i)
            for rng in ws[f"{col}2":f"{col}{ws.max_row}"]:
                for cell in rng: cell.number_format = "@"

        # valor
        col = _col_letter(val_idx + 1)
        for rng in ws[f"{col}2":f"{col}{ws.max_row}"]:
            for cell in rng: cell.number_format = number_format

        wb.save(path_xlsx)
    except Exception:
        pass

def run_pipe_taxas(input_file: Path, data_pagamento: str, saida_path: Path) -> dict:
    """
    Lê a planilha do usuário, trata e exporta FINAL e PENDENTES.
    - input_file: caminho do arquivo do usuário
    - data_pagamento: 'dd/mm/aaaa' (vem do launcher)
    - saida_path: nome-base sugerido (ex.: PipeTaxas_Final_AAAAMMDD.xlsx)
    """
    inp = Path(input_file)
    out = Path(saida_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    df_user = read_user_table(inp)
    df_norm = _normalize(df_user, data_pagamento)

    # ----- PENDENTES se faltar CNPJ ou QUALQUER DADO BANCÁRIO ou ENDEREÇO -----
    obrigatorios = [
        # identidade do favorecido
        "NOME_FAVORECIDO", "CADASTRO_FAVORECIDO",
        # dados bancários (fundo e favorecido)
        "CODIGO_BANCO_FUNDO", "AGENCIA_FUNDO", "CONTA_COM_DIGITO_FUNDO",
        "CODIGO_BANCO_FAVORECIDO", "AGENCIA_FAVORECIDO", "CONTA_COM_DIGITO_FAVORECIDO",
        # CNPJ do fundo
        "CNPJ_FUNDO",
        # endereço do favorecido
        "LOGRADOURO_FAVORECIDO", "CIDADE_FAVORECIDO", "CEP_FAVORECIDO",
    ]
    mask_ok = pd.Series(True, index=df_norm.index)
    for c in obrigatorios:
        mask_ok &= df_norm[c].astype(str).str.strip() != ""

    df_ok   = df_norm[mask_ok].copy()
    df_pend = df_norm[~mask_ok].copy()

    # Padrão: PipeTaxas_Tipo_AAAAMMDD
    data_hoje = dt.date.today().strftime('%Y%m%d')
    out_final = out.parent / f"PipeTaxas_Final_{data_hoje}.xlsx"
    out_pend  = out.parent / f"PipeTaxas_Pendentes_{data_hoje}.xlsx"

    _save_with_formats(df_ok, out_final, number_format="0.000000", date_format="dd/mm/yyyy", sheet_name="FINAL")
    if not df_pend.empty:
        _save_with_formats(df_pend, out_pend, number_format="0.000000", date_format="dd/mm/yyyy", sheet_name="PENDENTES")

    return {
        "saida_taxas_final": str(out_final),
        "saida_taxas_pendentes": (str(out_pend) if not df_pend.empty else ""),
        "qtd_total": int(len(df_norm)),
        "qtd_ok": int(len(df_ok)),
        "qtd_pendentes": int(len(df_pend)),
    }
