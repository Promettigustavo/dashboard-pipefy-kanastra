# liquidacao_core.py
import pandas as pd
import numpy as np
import re
import unicodedata
from pathlib import Path
import datetime as dt
import sys

# =====================
# Util para recursos (funciona em .py e .exe)
# =====================
def resource_path(rel: str) -> Path:
    """
    Retorna o caminho de um recurso tanto em execução normal quanto empacotada (.exe).
    Primeiro tenta ao lado do executável/script; se não achar, tenta a pasta temporária do PyInstaller.
    """
    exe_dir = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    cand = exe_dir / rel
    if cand.exists():
        return cand
    try:
        base_path = Path(sys._MEIPASS)  # type: ignore[attr-defined]
        return base_path / rel
    except Exception:
        return Path(__file__).parent / rel

# =====================
# Configuração
# =====================
CAMINHO_BASE = resource_path("Basedadosfundos.xlsx")  # manter ao lado do .exe ou embutir com --add-data

# Endereços fixos (escolha com ENDERECO_IDX: 0, 1 ou 2)
ENDERECO_IDX = 0
ENDERECOS_FIXOS = [
    {
        "LOGRADOURO_FAVORECIDO": "AVENIDA DOS VINHEDOS",
        "NUMERO_FAVORECIDO": "74",
        "COMPLEMENTO_FAVORECIDO": "SALA 805",
        "BAIRRO_FAVORECIDO": "JARDIM SUL",
        "CEP_FAVORECIDO": "38411851",
        "CIDADE_FAVORECIDO": "UBERLANDIA",
        "UF_FAVORECIDO": "MG",
    },
]

# =====================
# Utilidades
# =====================
def _ascii_upper(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().upper()

def canon_nome_fundo(valor: str) -> str:
    if pd.isna(valor):
        return ""
    s = str(valor)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper().strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    return s

def so_digitos(x: str) -> str:
    return re.sub(r"\D+", "", str(x)) if pd.notna(x) else ""

def zfill_or_empty(x: str, n: int) -> str:
    d = so_digitos(x)
    return d.zfill(n) if d else ""

def extrai_codigo_banco_inicio(s: str) -> str:
    if pd.isna(s):
        return ""
    m = re.match(r"\s*(\d{3,8})\b", str(s))
    return m.group(1) if m else ""

def parse_valor_to_string_with_4444(v: str) -> str:
    """
    Normaliza valor do usuário para string padrão, com 2 casas e sufixo '4444'.
    Ex.: '18645,97' -> '18645.974444'
    """
    s = str(v).strip()
    if s == "":
        return ""
    s = re.sub(r"[^\d,.-]", "", s)
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        num = float(s)
        return f"{num:.2f}" + "4444"
    except Exception:
        return str(v)

def clean_text(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and np.isnan(x):
        return ""
    sx = str(x).strip()
    if sx.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    return sx

# >>> Sanitização final para colunas textuais da saída
_REMOVE_CHARS_RE = re.compile(r"[.,/&\-\(\)\"'“”‘’]+")
def sanitize_text_out(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = _REMOVE_CHARS_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.upper()

def valida_data_pagamento(s: str) -> str:
    """
    Valida data no formato dd/mm/aaaa e não permite data anterior a hoje.
    Retorna a string normalizada dd/mm/aaaa.
    """
    s = re.sub(r"[.\- ]", "/", s.strip())
    hoje = dt.date.today()
    try:
        d = dt.datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        raise ValueError("Use o formato dd/mm/aaaa (ex.: 23/07/2025).")
    if d < hoje:
        raise ValueError(f"A data informada ({d.strftime('%d/%m/%Y')}) é anterior à data atual ({hoje.strftime('%d/%m/%Y')}).")
    return d.strftime("%d/%m/%Y")

# Conversor flexível para DATA -> datetime (Excel real)
def _parse_date_flex(val):
    """
    Aceita: 'dd/mm/aaaa', 'dd/mm/aa', 'aaaa-mm-dd' ou serial Excel.
    Retorna datetime.datetime ou None.
    """
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    # serial Excel
    try:
        sn = re.sub(r"[^0-9.\-]", "", s)
        if sn != "":
            serial = float(sn)
            base = dt.datetime(1899, 12, 30)
            d = base + dt.timedelta(days=serial)
            return dt.datetime(d.year, d.month, d.day)
    except Exception:
        pass
    return None

# =====================
# 1) Ler BANCO (todas as abas)
# =====================
def carregar_base_bancos(caminho_base: Path | str = CAMINHO_BASE) -> dict[str, pd.DataFrame]:
    return pd.read_excel(caminho_base, sheet_name=None, dtype=str, engine="openpyxl")

# =====================
# 2) Ler planilha do USUÁRIO (sem GUI)
# =====================
def read_user_file(path: Path | str) -> pd.DataFrame:
    import shutil, time
    p = Path(path)
    ext = p.suffix.lower()

    def _read_excel(q: Path) -> pd.DataFrame:
        eng = "openpyxl" if q.suffix.lower() == ".xlsx" else None
        return pd.read_excel(q, dtype=str, engine=eng)

    def _read_csv(q: Path) -> pd.DataFrame:
        return pd.read_csv(q, dtype=str, sep=None, engine="python")

    try:
        return _read_excel(p) if ext in (".xlsx", ".xls") else _read_csv(p)
    except PermissionError:
        tmp_dir = Path("_tmp_imports"); tmp_dir.mkdir(exist_ok=True)
        tmp_file = tmp_dir / f"copy_{p.stem}_{int(time.time())}{ext}"
        shutil.copy2(path, tmp_file)
        try:
            return _read_excel(tmp_file) if ext in (".xlsx", ".xls") else _read_csv(tmp_file)
        finally:
            try:
                tmp_file.unlink()
            except OSError:
                pass

# =====================
# 3) Normalizar planilha do USUÁRIO
# =====================
CANON = {
    "NOME DO FUNDO": "NOME_FUNDO",
    "RAZAO SOCIAL DO BENEFICIARIO": "RAZAO_BENEFICIARIO",
    "CNPJ": "CNPJ",
    "CPF": "CPF",
    "BANCO": "BANCO",
    "AGENCIA": "AGENCIA",
    "CONTA CORRENTE": "CONTA",
    "VALOR": "VALOR",
    "CODIGO DO BOLETO": "NUMERO_BOLETO",  # renomeia direto para NUMERO_BOLETO
    "CHAVE_PIX": "CHAVE_PIX",
}
def rename_to_canon(df: pd.DataFrame) -> pd.DataFrame:
    ren = {}
    for c in df.columns:
        key = _ascii_upper(c)
        if key in CANON:
            ren[c] = CANON[key]
    return df.rename(columns=ren)

def normaliza_usuario(df: pd.DataFrame) -> pd.DataFrame:
    d = rename_to_canon(df).copy()
    for need in [
        "NOME_FUNDO","RAZAO_BENEFICIARIO","CNPJ","CPF","BANCO","AGENCIA",
        "CONTA","VALOR","CHAVE_PIX","NUMERO_BOLETO"
    ]:
        if need not in d.columns:
            d[need] = ""
    d = d.astype(str)

    d["NOME_FUNDO"]         = d["NOME_FUNDO"].map(canon_nome_fundo)
    d["RAZAO_BENEFICIARIO"] = d["RAZAO_BENEFICIARIO"].map(clean_text)
    d["CNPJ"]               = d["CNPJ"].map(lambda x: so_digitos(x).zfill(14) if so_digitos(x) else "")
    d["CPF"]                = d["CPF"].map(lambda x: so_digitos(x).zfill(11) if so_digitos(x) else "")
    d["AGENCIA"]            = d["AGENCIA"].map(lambda x: zfill_or_empty(x, 4))
    d["CONTA"]              = d["CONTA"].map(lambda x: zfill_or_empty(x, 8))
    d["COD_BANCO"]          = d["BANCO"].map(extrai_codigo_banco_inicio)
    d["VALOR"]              = d["VALOR"].map(parse_valor_to_string_with_4444)
    d["CHAVE_PIX"]          = d["CHAVE_PIX"].map(clean_text)
    d["NUMERO_BOLETO"]      = d["NUMERO_BOLETO"].map(so_digitos)

    for col in d.columns:
        d[col] = d[col].fillna("").astype(str).str.strip()
    return d

# =====================
# 4) Normalizar BANCO
# =====================
BANK_CANON = {
    "RAZAO SOCIAL FUNDO": "NOME_FUNDO",
    "NOME DO FUNDO": "NOME_FUNDO",
    "NOME_FUNDO": "NOME_FUNDO",
    "CNPJ FUNDO": "CNPJ_FUNDO",
    "CNPJ DO FUNDO": "CNPJ_FUNDO",
    "CNPJ_FUNDO": "CNPJ_FUNDO",
    "BANCO FUNDO": "BANCO",
    "BANCO": "BANCO",
    "AGENCIA FUNDO": "AGENCIA",
    "AGÊNCIA FUNDO": "AGENCIA",
    "AGENCIA": "AGENCIA",
    "AGÊNCIA": "AGENCIA",
    "CONTA CORRENTE FUNDO": "CONTA",
    "CONTA CORRENTE": "CONTA",
    "CODIGO DO BANCO": "CODIGO_BANCO",
    "CODIGO BANCO": "CODIGO_BANCO",
    "CÓDIGO BANCO": "CODIGO_BANCO",
    "COD_BANCO": "CODIGO_BANCO",
    "DV_CONTA": "DV_CONTA",
    "DV": "DV_CONTA",
}
def bank_rename_to_canon(df: pd.DataFrame) -> pd.DataFrame:
    def _canon_key(s: str) -> str:
        s = unicodedata.normalize("NFKD", str(s))
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.upper().strip().replace("_", " ")
        s = re.sub(r"\s+", " ", s)
        return s

    ren = {}
    for col in df.columns:
        key = _canon_key(col)
        if key in BANK_CANON:
            ren[col] = BANK_CANON[key]
    return df.rename(columns=ren)

def normaliza_banco(abas_dict: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = []
    for aba, df in abas_dict.items():
        if df is None or df.empty:
            continue
        d = df.copy()
        d["ABA_ORIGEM"] = aba
        frames.append(d)
    if not frames:
        return pd.DataFrame()

    base = pd.concat(frames, ignore_index=True).astype(str)
    base.columns = [str(c).strip() for c in base.columns]
    base = bank_rename_to_canon(base)

    for col in ["NOME_FUNDO","CNPJ_FUNDO","CODIGO_BANCO","BANCO","AGENCIA","CONTA","DV_CONTA"]:
        if col not in base.columns:
            base[col] = ""

    base["NOME_FUNDO"]  = base["NOME_FUNDO"].fillna("").astype(str).map(canon_nome_fundo)
    base["CNPJ_FUNDO"]  = base["CNPJ_FUNDO"].map(lambda x: so_digitos(x).zfill(14) if so_digitos(x) else "")
    base["AGENCIA"]     = base["AGENCIA"].map(lambda x: so_digitos(x).zfill(4) if so_digitos(x) else "")
    base["CONTA"]       = base["CONTA"].map(lambda x: so_digitos(x))
    base["DV_CONTA"]    = base["DV_CONTA"].map(lambda x: so_digitos(x)[-1:] if so_digitos(x) else "")

    base["CONTA_COM_DIGITO"] = base.apply(
        lambda r: f"{r['CONTA']}-{r['DV_CONTA']}" if r.get("DV_CONTA","") else str(r.get("CONTA","")),
        axis=1
    )
    base["CHAVE_FUNDO"] = base["NOME_FUNDO"]
    base = base.drop_duplicates(subset=["CHAVE_FUNDO"], keep="first")

    return base[["CHAVE_FUNDO","CNPJ_FUNDO","CODIGO_BANCO","AGENCIA","CONTA_COM_DIGITO"]].copy()

def prepara_usuario_para_merge(df_user_norm: pd.DataFrame) -> pd.DataFrame:
    d = df_user_norm.copy()
    if "NOME_FUNDO" not in d.columns:
        d["NOME_FUNDO"] = ""
    d["NOME_FUNDO"] = d["NOME_FUNDO"].map(canon_nome_fundo)
    d["CHAVE_FUNDO"] = d["NOME_FUNDO"]
    return d

# =====================
# 5) Geração da SAÍDA e dos NÃO IMPORTADOS
# =====================
def gerar_saida_para_time(df_user_norm: pd.DataFrame,
                          abas_dict: dict[str, pd.DataFrame],
                          data_pagamento: str,
                          saida_path: Path | str | None = None) -> Path:
    base_norm = normaliza_banco(abas_dict).rename(columns={
        "CNPJ_FUNDO": "CNPJ_FUNDO_BANCO",
        "CODIGO_BANCO": "CODIGO_BANCO_FUNDO",
        "AGENCIA": "AGENCIA_FUNDO",
        "CONTA_COM_DIGITO": "CONTA_COM_DIGITO_FUNDO",
    })

    user_norm = prepara_usuario_para_merge(df_user_norm)

    # indicador para separar não importados
    merged = user_norm.merge(base_norm, on="CHAVE_FUNDO", how="left", indicator=True)

    # helper: garante Series mesmo que a coluna não exista
    def _col_or_blank(df: pd.DataFrame, col: str) -> pd.Series:
        if col in df.columns:
            return df[col].astype(str)
        return pd.Series([""] * len(df), index=df.index, dtype=str)

    def _cadastro_fav(row) -> str:
        cnpj = so_digitos(row.get("CNPJ", ""))[:14]
        cpf  = so_digitos(row.get("CPF", ""))[:11]
        if cnpj:
            return cnpj.zfill(14)
        if cpf:
            return cpf.zfill(11)
        return ""

    saida = pd.DataFrame({
        "CNPJ_FUNDO":                  _col_or_blank(merged, "CNPJ_FUNDO_BANCO").map(clean_text),
        "CODIGO_BANCO_FUNDO":          _col_or_blank(merged, "CODIGO_BANCO_FUNDO").map(clean_text),
        "AGENCIA_FUNDO":               _col_or_blank(merged, "AGENCIA_FUNDO").map(clean_text),
        "CONTA_COM_DIGITO_FUNDO":      _col_or_blank(merged, "CONTA_COM_DIGITO_FUNDO").map(clean_text),
        "NOME_FAVORECIDO":             _col_or_blank(merged, "RAZAO_BENEFICIARIO").map(clean_text),
        "CADASTRO_FAVORECIDO":         merged.apply(_cadastro_fav, axis=1).map(clean_text),
        "CODIGO_BANCO_FAVORECIDO":     _col_or_blank(merged, "COD_BANCO").map(clean_text),
        "AGENCIA_FAVORECIDO":          _col_or_blank(merged, "AGENCIA").map(clean_text),
        "CONTA_COM_DIGITO_FAVORECIDO": _col_or_blank(merged, "CONTA").map(clean_text),
        "VALOR_FAVORECIDO":            _col_or_blank(merged, "VALOR").map(clean_text),
        "LOGRADOURO_FAVORECIDO":       ENDERECOS_FIXOS[ENDERECO_IDX]["LOGRADOURO_FAVORECIDO"],
        "NUMERO_FAVORECIDO":           ENDERECOS_FIXOS[ENDERECO_IDX]["NUMERO_FAVORECIDO"],
        "COMPLEMENTO_FAVORECIDO":      ENDERECOS_FIXOS[ENDERECO_IDX]["COMPLEMENTO_FAVORECIDO"],
        "BAIRRO_FAVORECIDO":           ENDERECOS_FIXOS[ENDERECO_IDX]["BAIRRO_FAVORECIDO"],
        "CEP_FAVORECIDO":              so_digitos(ENDERECOS_FIXOS[ENDERECO_IDX]["CEP_FAVORECIDO"]).zfill(8),
        "CIDADE_FAVORECIDO":           ENDERECOS_FIXOS[ENDERECO_IDX]["CIDADE_FAVORECIDO"],
        "UF_FAVORECIDO":               ENDERECOS_FIXOS[ENDERECO_IDX]["UF_FAVORECIDO"],
        "DATA_PAGAMENTO":              data_pagamento,
        "NUMERO_BOLETO":               _col_or_blank(merged, "NUMERO_BOLETO").map(so_digitos),
        "CHAVE_PIX":                   _col_or_blank(merged, "CHAVE_PIX").map(clean_text),
        "_merge":                      merged["_merge"].astype(str),
    })

    # manter na principal apenas os que deram match no banco
    mask_principal = saida["_merge"] == "both"

    # AGENCIA_FAVORECIDO: regra especial para 5 dígitos
    def _ajusta_agencia(valor: str) -> str:
        v = so_digitos(valor)
        if len(v) == 5:
            if v.startswith("0"):
                return v[1:]   # remove o 1º (mantém os 4 últimos)
            return v[:4]       # remove o último (mantém os 4 primeiros)
        elif len(v) in (3, 4):
            return v
        return v
    saida["AGENCIA_FAVORECIDO"] = saida["AGENCIA_FAVORECIDO"].map(_ajusta_agencia)

    # sanitização final das colunas textuais
    for col in ["NOME_FAVORECIDO","LOGRADOURO_FAVORECIDO","COMPLEMENTO_FAVORECIDO",
                "BAIRRO_FAVORECIDO","CIDADE_FAVORECIDO"]:
        saida[col] = saida[col].map(sanitize_text_out)

    # Ordem final (mantendo boleto/PIX ao fim)
    cols_final = [
        "CNPJ_FUNDO","CODIGO_BANCO_FUNDO","AGENCIA_FUNDO","CONTA_COM_DIGITO_FUNDO",
        "NOME_FAVORECIDO","CADASTRO_FAVORECIDO","CODIGO_BANCO_FAVORECIDO",
        "AGENCIA_FAVORECIDO","CONTA_COM_DIGITO_FAVORECIDO","VALOR_FAVORECIDO",
        "LOGRADOURO_FAVORECIDO","NUMERO_FAVORECIDO","COMPLEMENTO_FAVORECIDO",
        "BAIRRO_FAVORECIDO","CEP_FAVORECIDO","CIDADE_FAVORECIDO","UF_FAVORECIDO",
        "DATA_PAGAMENTO","NUMERO_BOLETO","CHAVE_PIX","_merge"
    ]
    saida = saida[cols_final].fillna("").astype(str)

    # Caminhos com padrão: Processo_Tipo_AAAAMMDD
    saida_path = Path(saida_path) if saida_path is not None else Path("Liquidacao_Remessa.xlsx")
    out_dir = Path(saida_path).parent
    data_hoje = dt.date.today().strftime('%Y%m%d')
    nao_imp_path = out_dir / f"Liquidacao_NaoImportados_{data_hoje}.xlsx"

    # Conversor para VALOR_FAVORECIDO -> float com 6 casas
    def _valor_str_to_float(v: str):
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", ".")
        s = re.sub(r"[^0-9\.\-]", "", s)
        try:
            return float(s)
        except Exception:
            return None

    # ---------- Gravar planilha principal (somente matched) ----------
    principal = saida[mask_principal].drop(columns=["_merge"]).copy()

    try:
        with pd.ExcelWriter(saida_path, engine="openpyxl") as writer:
            sheet = "SAIDA"
            principal.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]

            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers

            n_rows, n_cols = principal.shape
            cols_list = list(principal.columns)
            idx_valor = cols_list.index("VALOR_FAVORECIDO") + 1  # 1-based
            idx_data  = cols_list.index("DATA_PAGAMENTO") + 1    # 1-based

            # Cabeçalho texto
            for c in range(1, n_cols + 1):
                ws.cell(row=1, column=c).number_format = "@"

            # Dados: tudo texto, exceto VALOR_FAVORECIDO (número). DATA_PAGAMENTO agora forçada como TEXTO.
            for r in range(2, n_rows + 2):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    if c == idx_valor:
                        f = _valor_str_to_float(cell.value)
                        cell.value = f if f is not None else None
                        cell.number_format = "0.000000"
                    elif c == idx_data:
                        # >>> ALTERADO: manter como TEXTO <<<
                        cell.value = "" if cell.value is None else str(cell.value)
                        cell.number_format = "@"
                    else:
                        cell.value = "" if cell.value is None else str(cell.value)
                        cell.number_format = "@"

            # Larguras
            for idx, col in enumerate(cols_list, start=1):
                if idx == idx_valor:
                    ws.column_dimensions[get_column_letter(idx)].width = 18
                elif idx == idx_data:
                    ws.column_dimensions[get_column_letter(idx)].width = 12
                else:
                    maxlen = max(len(str(col)), *(len(s) for s in principal[col].astype(str).values)) + 2
                    ws.column_dimensions[get_column_letter(idx)].width = min(maxlen, 60)
    except Exception:
        principal.to_excel(saida_path, index=False)

    # ---------- Gravar planilha dos NÃO IMPORTADOS ----------
    nao_imp = saida[~mask_principal].drop(columns=["_merge"]).copy()
    nao_imp = nao_imp[[c for c in cols_final if c != "_merge"]].reset_index(drop=True)

    # Reabrimos um merge rápido para pegar nomes dos não importados:
    base_norm_tmp = normaliza_banco(abas_dict=carregar_base_bancos()).rename(columns={
        "CNPJ_FUNDO": "CNPJ_FUNDO_BANCO",
        "CODIGO_BANCO": "CODIGO_BANCO_FUNDO",
        "AGENCIA": "AGENCIA_FUNDO",
        "CONTA_COM_DIGITO": "CONTA_COM_DIGITO_FUNDO",
    })
    user_norm_tmp = prepara_usuario_para_merge(df_user_norm)
    merged_tmp = user_norm_tmp.merge(base_norm_tmp, on="CHAVE_FUNDO", how="left", indicator=True)
    mask_principal_tmp = merged_tmp["_merge"] == "both"
    nomes_fundos = merged_tmp.loc[~mask_principal_tmp, "NOME_FUNDO"].fillna("").astype(str).tolist()

    if not nao_imp.empty:
        # Nome do fundo na frente
        nao_imp["CNPJ_FUNDO"] = [
            f"{nome} - FUNDO NAO CADASTRADO NO BANCO DE DADOS" if nome
            else "FUNDO NAO CADASTRADO NO BANCO DE DADOS"
            for nome in nomes_fundos
        ]
        nao_imp["CODIGO_BANCO_FUNDO"] = ""
        nao_imp["AGENCIA_FUNDO"] = ""
        nao_imp["CONTA_COM_DIGITO_FUNDO"] = ""

        try:
            with pd.ExcelWriter(nao_imp_path, engine="openpyxl") as writer:
                sheet = "Fundosnaoimportados"
                nao_imp.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]

                from openpyxl.utils import get_column_letter
                from openpyxl.styles import numbers

                n_rows, n_cols = nao_imp.shape
                cols_list = list(nao_imp.columns)
                idx_valor = cols_list.index("VALOR_FAVORECIDO") + 1
                idx_data  = cols_list.index("DATA_PAGAMENTO") + 1

                # Cabeçalho texto
                for c in range(1, n_cols + 1):
                    ws.cell(row=1, column=c).number_format = "@"

                # Dados (DATA_PAGAMENTO agora TEXTO)
                for r in range(2, n_rows + 2):
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        if c == idx_valor:
                            f = _valor_str_to_float(cell.value)
                            cell.value = f if f is not None else None
                            cell.number_format = "0.000000"
                        elif c == idx_data:
                            # >>> ALTERADO: manter como TEXTO <<<
                            cell.value = "" if cell.value is None else str(cell.value)
                            cell.number_format = "@"
                        else:
                            cell.value = "" if cell.value is None else str(cell.value)
                            cell.number_format = "@"

                # Larguras
                for idx, col in enumerate(cols_list, start=1):
                    if idx == idx_valor:
                        ws.column_dimensions[get_column_letter(idx)].width = 18
                    elif idx == idx_data:
                        ws.column_dimensions[get_column_letter(idx)].width = 12
                    else:
                        maxlen = max(len(str(col)), *(len(s) for s in nao_imp[col].astype(str).values)) + 2
                        ws.column_dimensions[get_column_letter(idx)].width = min(maxlen, 60)
        except Exception:
            nao_imp.to_excel(nao_imp_path, index=False)

    print(f"OK Planilha principal: {Path(saida_path).resolve()} (linhas: {len(principal)})")
    if not nao_imp.empty:
        print(f"OK Fundos não importados: {nao_imp_path.resolve()} (linhas: {len(nao_imp)})")
    else:
        print("OK Todos os fundos foram importados. Nenhuma planilha de 'não importados' foi gerada.")
    return Path(saida_path)

# =====================
# Função de orquestração sem GUI
# =====================
def run_liquidacao(input_file: str | Path,
                   data_pagamento: str,
                   saida_path: str | Path | None = None) -> dict:
    """
    Executa todo o fluxo da liquidação sem GUI.
    - input_file: caminho da planilha do usuário (.xlsx/.xls/.csv/.txt)
    - data_pagamento: dd/mm/aaaa (não pode ser anterior a hoje)
    - saida_path: caminho de saída (opcional). Se None, gera Remessa_YYYYMMDD.xlsx ao lado do input_file.
    """
    data_ok = valida_data_pagamento(data_pagamento)

    # 1) Carrega bases
    abas = carregar_base_bancos()
    df_user_raw = read_user_file(input_file)
    df_user_raw.columns = [str(c).strip().upper() for c in df_user_raw.columns]
    df_user_norm = normaliza_usuario(df_user_raw)

    # 2) Nome de saída padrão: Processo_Tipo_AAAAMMDD
    if saida_path is None:
        data_hoje = dt.date.today().strftime('%Y%m%d')
        saida_path = Path(input_file).parent / f"Liquidacao_Remessa_{data_hoje}.xlsx"

    # 3) Geração
    out_path = gerar_saida_para_time(df_user_norm, abas, data_pagamento=data_ok, saida_path=saida_path)

    # 4) Estatísticas simples
    base_norm_tmp = normaliza_banco(abas).rename(columns={
        "CNPJ_FUNDO": "CNPJ_FUNDO_BANCO",
        "CODIGO_BANCO": "CODIGO_BANCO_FUNDO",
        "AGENCIA": "AGENCIA_FUNDO",
        "CONTA_COM_DIGITO": "CONTA_COM_DIGITO_FUNDO",
    })
    user_norm_tmp = prepara_usuario_para_merge(df_user_norm)
    merged_tmp = user_norm_tmp.merge(base_norm_tmp, on="CHAVE_FUNDO", how="left", indicator=True)
    nao_importados = merged_tmp.loc[merged_tmp["_merge"] != "both", "NOME_FUNDO"].dropna().unique().tolist()

    return {
        "saida_principal": str(Path(out_path).resolve()),
        "nao_importados": str((Path(out_path).parent / "Fundosnãoimportados.xlsx").resolve()),
        "qtd_total": int(len(df_user_norm)),
        "qtd_nao_importados": int(len(nao_importados)),
        "lista_nao_importados": nao_importados,
    }

# =====================
# CLI (opcional)
# =====================
if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Pipe de Liquidação (sem GUI)")
    ap.add_argument("--input", required=True, help="Planilha do usuário (.xlsx/.xls/.csv/.txt)")
    ap.add_argument("--data", required=True, help="Data de pagamento no formato dd/mm/aaaa (não pode ser anterior a hoje)")
    ap.add_argument("--out", help="Caminho do arquivo de saída (.xlsx). Opcional.")
    args = ap.parse_args()

    info = run_liquidacao(args.input, args.data, args.out)
    print("\nResumo:")
    for k, v in info.items():
        if k == "lista_nao_importados":
            print(f"- {k}: {', '.join(v) if v else 'Nenhum'}")
        else:
            print(f"- {k}: {v}")
